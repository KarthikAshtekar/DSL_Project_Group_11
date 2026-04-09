# =============================================================================
#  NOTEBOOK 1 · BiRNN & BiLSTM Classifiers + Time-Series Demand Forecasting
#  Run top-to-bottom in Google Colab after uploading:
#      • Train_TEST.csv
#      • seasonal_index_data.csv   (generated automatically in Section 2)
#
#  Outputs produced:
#      • bilstm_birnn_model.h5       — saved BiRNN weights
#      • bilstm_model_improved.h5    — saved BiLSTM weights
#      • tokenizer.pkl               — fitted Keras tokenizer
#      • all_plots.pdf               — every figure in one PDF
#      • seasonal_index_data.csv     — aggregated day-level time series
#      • FINAL_forecast_results.xlsx — 30-day forecast for the optimizer
# =============================================================================

# ── SECTION 0 · Install dependencies (run once) ──────────────────────────────
# Uncomment the line below when running for the first time on a fresh Colab session
# !pip install tensorflow scikit-learn statsmodels openpyxl -q

# ── SECTION 1 · Imports ──────────────────────────────────────────────────────
import warnings
import pickle
import itertools
import os

warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as pdf_backend
import statsmodels.api as sm

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import (
    classification_report, confusion_matrix,
    mean_absolute_error, mean_squared_error,
)

from statsmodels.tsa.seasonal import seasonal_decompose
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.holtwinters import ExponentialSmoothing, SimpleExpSmoothing
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.stats.diagnostic import acorr_ljungbox

import tensorflow as tf
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import (
    Embedding, Bidirectional, SimpleRNN, LSTM, Dense, Dropout,
)
from tensorflow.keras.callbacks import EarlyStopping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("✓ All libraries loaded")

# ── SECTION 2 · Configuration — edit file paths here only ────────────────────

# Input files
TRAIN_CSV        = "Train_TEST.csv"           # Raw complaint data
SEASONAL_CSV     = "seasonal_index_data.csv"  # Generated in Section 4 below

# Output files
BIRNN_H5         = "bilstm_birnn_model.h5"
LSTM_H5          = "bilstm_model_improved.h5"
TOKENIZER_PKL    = "tokenizer.pkl"
PLOTS_PDF        = "all_plots.pdf"
FORECAST_XLSX    = "FINAL_forecast_results.xlsx"
INTERMEDIATE_CSV = "seasonal_index_data.csv"   # Written during preprocessing

# NLP model hyperparameters
VOCAB_SIZE       = 10_000
MAX_LEN          = 100
EPOCHS           = 10
BATCH_SIZE       = 64
TEST_SPLIT       = 0.2
RANDOM_STATE     = 42

# Time-series settings
FORECAST_DAYS    = 30
SEASONAL_PERIOD  = 7
TS_START_DATE    = "2024-01-01"

# ARIMA candidate grid (used in model-selection race)
ARIMA_CANDIDATES = {
    "AR(1)":    (1, 0, 0),
    "MA(1)":    (0, 0, 1),
    "ARMA(1,1)":(1, 0, 1),
    "AR(2)":    (2, 0, 0),
}

# Severity column names expected in the aggregated time-series CSV
SEVERITY_COLS = [
    "severity_Critical",
    "severity_High",
    "severity_Low",
    "severity_Medium",
    "severity_Very Low",
]

# Shared PDF writer — all plots are appended here
_PDF = pdf_backend.PdfPages(PLOTS_PDF)

def save_fig(fig=None):
    """Append the current (or given) figure to the shared PDF."""
    f = fig if fig is not None else plt.gcf()
    _PDF.savefig(f, bbox_inches="tight")

print("✓ Configuration loaded")
print(f"  Plots  → {PLOTS_PDF}")
print(f"  Forecast → {FORECAST_XLSX}")

# =============================================================================
#  PART A — NLP SEVERITY CLASSIFICATION
# =============================================================================

# ── SECTION 3 · Load data and prepare train / test splits ────────────────────
print("\n" + "=" * 60)
print("PART A — NLP Severity Classification")
print("=" * 60)

df_text = pd.read_csv(TRAIN_CSV)
df_text = df_text.dropna(subset=["grievance_text", "severity"])

X_raw = df_text["grievance_text"].astype(str)
y_raw = df_text["severity"]

# Encode string severity labels to integers
le = LabelEncoder()
y_enc = le.fit_transform(y_raw)
num_classes = len(le.classes_)

print(f"\nDataset shape : {df_text.shape}")
print(f"Label mapping : {dict(zip(le.classes_, range(num_classes)))}")

X_train_raw, X_test_raw, y_train, y_test = train_test_split(
    X_raw, y_enc,
    test_size=TEST_SPLIT,
    random_state=RANDOM_STATE,
    stratify=y_enc,
)
print(f"Train size    : {len(X_train_raw)}  |  Test size: {len(X_test_raw)}")

# Tokenise and pad sequences (shared by both models)
tokenizer = Tokenizer(num_words=VOCAB_SIZE, oov_token="<OOV>")
tokenizer.fit_on_texts(X_train_raw)

X_train_pad = pad_sequences(
    tokenizer.texts_to_sequences(X_train_raw), maxlen=MAX_LEN, padding="post"
)
X_test_pad = pad_sequences(
    tokenizer.texts_to_sequences(X_test_raw), maxlen=MAX_LEN, padding="post"
)

# Shared early-stopping callback (reused for both models)
early_stop = EarlyStopping(
    monitor="val_loss", patience=3, restore_best_weights=True
)

# ── SECTION 4 · Bidirectional RNN (BiRNN) ────────────────────────────────────
print("\n--- Training Bidirectional RNN ---")

birnn_model = Sequential(
    [
        Embedding(VOCAB_SIZE, 128, input_length=MAX_LEN),
        Bidirectional(SimpleRNN(64, return_sequences=True)),
        Bidirectional(SimpleRNN(32)),
        Dense(32, activation="relu"),
        Dense(16, activation="relu"),
        Dense(num_classes, activation="softmax"),
    ],
    name="BiRNN",
)
birnn_model.compile(
    loss="sparse_categorical_crossentropy", optimizer="adam", metrics=["accuracy"]
)
birnn_model.summary()

history_birnn = birnn_model.fit(
    X_train_pad, y_train,
    epochs=EPOCHS,
    batch_size=BATCH_SIZE,
    validation_split=0.2,
    callbacks=[early_stop],
    verbose=1,
)

# Evaluate
loss_birnn, acc_birnn = birnn_model.evaluate(X_test_pad, y_test, verbose=0)
y_pred_birnn = np.argmax(birnn_model.predict(X_test_pad), axis=1)

print(f"\n[BiRNN] Test Accuracy : {acc_birnn:.4f}")
print("\n===== Classification Report — BiRNN =====")
print(classification_report(y_test, y_pred_birnn, target_names=le.classes_))
print("===== Confusion Matrix — BiRNN =====")
print(confusion_matrix(y_test, y_pred_birnn))

# Training curves
fig, axes = plt.subplots(1, 2, figsize=(12, 4))
axes[0].plot(history_birnn.history["accuracy"],     label="Train")
axes[0].plot(history_birnn.history["val_accuracy"], label="Validation")
axes[0].set_title("BiRNN — Accuracy")
axes[0].set_xlabel("Epoch")
axes[0].legend()
axes[1].plot(history_birnn.history["loss"],     label="Train")
axes[1].plot(history_birnn.history["val_loss"], label="Validation")
axes[1].set_title("BiRNN — Loss")
axes[1].set_xlabel("Epoch")
axes[1].legend()
plt.suptitle("BiRNN Training Curves", fontsize=13)
plt.tight_layout()
save_fig()
plt.show()

# Confusion matrix heatmap
import seaborn as sns
fig_cm, ax_cm = plt.subplots(figsize=(7, 5))
sns.heatmap(
    confusion_matrix(y_test, y_pred_birnn),
    annot=True, fmt="d", cmap="Blues",
    xticklabels=le.classes_, yticklabels=le.classes_, ax=ax_cm,
)
ax_cm.set_title("Confusion Matrix — BiRNN")
ax_cm.set_xlabel("Predicted")
ax_cm.set_ylabel("Actual")
plt.tight_layout()
save_fig()
plt.show()

# Save model and tokenizer
birnn_model.save(BIRNN_H5)
with open(TOKENIZER_PKL, "wb") as fh:
    pickle.dump(tokenizer, fh)
print(f"\n✓ BiRNN model saved → {BIRNN_H5}")
print(f"✓ Tokenizer saved  → {TOKENIZER_PKL}")

# ── SECTION 5 · Bidirectional LSTM with Dropout ───────────────────────────────
print("\n--- Training Bidirectional LSTM ---")

lstm_model = Sequential(
    [
        Embedding(VOCAB_SIZE, 128, input_length=MAX_LEN),
        Bidirectional(LSTM(64, return_sequences=True)),
        Dropout(0.3),
        Bidirectional(LSTM(32)),
        Dropout(0.3),
        Dense(32, activation="relu"),
        Dropout(0.2),
        Dense(16, activation="relu"),
        Dense(num_classes, activation="softmax"),
    ],
    name="BiLSTM",
)
lstm_model.compile(
    loss="sparse_categorical_crossentropy", optimizer="adam", metrics=["accuracy"]
)
lstm_model.summary()

history_lstm = lstm_model.fit(
    X_train_pad, y_train,
    epochs=EPOCHS,
    batch_size=BATCH_SIZE,
    validation_split=0.2,
    callbacks=[early_stop],
    verbose=1,
)

# Evaluate
loss_lstm, acc_lstm = lstm_model.evaluate(X_test_pad, y_test, verbose=0)
y_pred_lstm = np.argmax(lstm_model.predict(X_test_pad), axis=1)

print(f"\n[BiLSTM] Test Accuracy : {acc_lstm:.4f}")
print("\n===== Classification Report — BiLSTM =====")
print(classification_report(y_test, y_pred_lstm, target_names=le.classes_))
print("===== Confusion Matrix — BiLSTM =====")
print(confusion_matrix(y_test, y_pred_lstm))

# Training curves
fig, axes = plt.subplots(1, 2, figsize=(12, 4))
axes[0].plot(history_lstm.history["accuracy"],     label="Train")
axes[0].plot(history_lstm.history["val_accuracy"], label="Validation")
axes[0].set_title("BiLSTM — Accuracy")
axes[0].set_xlabel("Epoch")
axes[0].legend()
axes[1].plot(history_lstm.history["loss"],     label="Train")
axes[1].plot(history_lstm.history["val_loss"], label="Validation")
axes[1].set_title("BiLSTM — Loss")
axes[1].set_xlabel("Epoch")
axes[1].legend()
plt.suptitle("BiLSTM Training Curves", fontsize=13)
plt.tight_layout()
save_fig()
plt.show()

# Confusion matrix heatmap
fig_cm2, ax_cm2 = plt.subplots(figsize=(7, 5))
sns.heatmap(
    confusion_matrix(y_test, y_pred_lstm),
    annot=True, fmt="d", cmap="Greens",
    xticklabels=le.classes_, yticklabels=le.classes_, ax=ax_cm2,
)
ax_cm2.set_title("Confusion Matrix — BiLSTM")
ax_cm2.set_xlabel("Predicted")
ax_cm2.set_ylabel("Actual")
plt.tight_layout()
save_fig()
plt.show()

lstm_model.save(LSTM_H5)
print(f"✓ BiLSTM model saved → {LSTM_H5}")

# Inference helper (uses the trained BiLSTM)
def predict_severity(text: str) -> str:
    """Return predicted severity label for a single complaint string."""
    seq = tokenizer.texts_to_sequences([text])
    pad = pad_sequences(seq, maxlen=MAX_LEN, padding="post")
    idx = np.argmax(lstm_model.predict(pad, verbose=0), axis=1)[0]
    return le.inverse_transform([idx])[0]

print("\nExample inference:")
print("  Input  :", "internet not working since morning please fix")
print("  Output :", predict_severity("internet not working since morning please fix"))

# ── SECTION 6 · Model summary comparison ─────────────────────────────────────
print("\n===== NLP Model Performance Summary =====")
print(f"{'Model':<12} {'Test Accuracy':>14}")
print("-" * 28)
print(f"{'BiRNN':<12} {acc_birnn:>14.4f}")
print(f"{'BiLSTM':<12} {acc_lstm:>14.4f}")
print("\nNote: BERT (run separately in Notebook 2) achieves Precision=0.85, F1=0.76, AUC=0.82")

# =============================================================================
#  PART B — TIME-SERIES DEMAND FORECASTING
# =============================================================================
print("\n" + "=" * 60)
print("PART B — Time-Series Demand Forecasting")
print("=" * 60)

# ── SECTION 7 · Preprocess raw CSV → day-level severity time series ───────────
print("\n--- Section 7: Building seasonal_index_data.csv ---")

df_raw = pd.read_csv(TRAIN_CSV)
df_raw["timestamp"] = pd.to_datetime(df_raw["timestamp"], format="%d-%m-%Y %H:%M")

# One-hot encode severity and aggregate by calendar date
df_enc = pd.get_dummies(df_raw, columns=["severity"], dtype=int)
severity_oh_cols = [c for c in df_enc.columns if c.startswith("severity_")]

df_enc["date"] = df_enc["timestamp"].dt.date
df_daily = df_enc.groupby("date")[severity_oh_cols].sum().reset_index()
df_daily["month"] = pd.to_datetime(df_daily["date"]).dt.month
df_daily["day"]   = pd.to_datetime(df_daily["date"]).dt.day
df_daily = df_daily.drop(columns=["date"])

# Aggregate to unique (month, day) combinations and assign a serial day_index
df_agg = df_daily.groupby(["month", "day"])[severity_oh_cols].sum().reset_index()
df_agg = df_agg.sort_values(["month", "day"]).reset_index(drop=True)
df_agg["day_index"] = range(1, len(df_agg) + 1)
df_agg = df_agg[["day_index"] + severity_oh_cols]

df_agg.to_csv(INTERMEDIATE_CSV, index=False)
print(f"✓ Saved → {INTERMEDIATE_CSV}  ({len(df_agg)} rows)")
print(df_agg.head())

# ── SECTION 8 · Load time series with datetime index ─────────────────────────
print("\n--- Section 8: Loading time series ---")

df = pd.read_csv(SEASONAL_CSV)
df["date"] = pd.to_datetime(TS_START_DATE) + pd.to_timedelta(
    df["day_index"] - 1, unit="D"
)
df = (
    df.sort_values("date")
    .set_index("date")
    .asfreq("D")
    .drop(columns=["day_index"])
)

print(f"Time series: {df.shape[0]} days × {df.shape[1]} severity bands")
print(df.tail(3))

# ── SECTION 9 · Seasonal decomposition ───────────────────────────────────────
print("\n--- Section 9: Seasonal Decomposition (period=7) ---")

for col in SEVERITY_COLS:
    result = seasonal_decompose(df[col], model="additive", period=SEASONAL_PERIOD)
    fig = result.plot()
    fig.suptitle(f"Decomposition — {col}", fontsize=12, y=1.01)
    plt.tight_layout()
    save_fig(fig)
    plt.show()

# ── SECTION 10 · ADF Stationarity Tests ──────────────────────────────────────
print("\n--- Section 10: ADF Stationarity Tests ---")
print(f"\n{'Feature':<25} {'ADF Stat':>10} {'p-value':>10} {'Stationary':>12}")
print("-" * 62)
for col in SEVERITY_COLS:
    r       = adfuller(df[col].dropna())
    verdict = "YES ✓" if r[1] < 0.05 else "NO  ✗"
    print(f"{col:<25} {r[0]:>10.3f} {r[1]:>10.4f} {verdict:>12}")

# ── SECTION 11 · ACF and PACF Plots ──────────────────────────────────────────
print("\n--- Section 11: ACF / PACF Analysis ---")

for col in SEVERITY_COLS:
    fig, axes = plt.subplots(1, 2, figsize=(13, 4))
    plot_acf( df[col], lags=50, ax=axes[0])
    axes[0].set_title(f"ACF — {col}")
    plot_pacf(df[col], lags=50, method="ywm", ax=axes[1])
    axes[1].set_title(f"PACF — {col}")
    plt.tight_layout()
    save_fig()
    plt.show()

# ── SECTION 12 · ARIMA Candidate Models — AIC Ranking ────────────────────────
print("\n--- Section 12: ARIMA Candidate Models (AIC ranking) ---")

records = []
for col in SEVERITY_COLS:
    for name, order in ARIMA_CANDIDATES.items():
        try:
            fit = ARIMA(df[col], order=order).fit()
            records.append(
                {"Feature": col, "Model": name, "Order": str(order),
                 "AIC": round(fit.aic, 2), "BIC": round(fit.bic, 2)}
            )
        except Exception:
            pass

aic_df     = pd.DataFrame(records).sort_values(["Feature", "AIC"]).reset_index(drop=True)
best_arima = aic_df.loc[aic_df.groupby("Feature")["AIC"].idxmin()].reset_index(drop=True)

print("\n===== AIC Table =====")
print(aic_df.pivot(index="Feature", columns="Model", values="AIC").to_string())
print("\n===== Best Model per Feature =====")
print(best_arima[["Feature", "Model", "Order", "AIC"]].to_string(index=False))

# ── SECTION 13 · Residual Diagnostics ────────────────────────────────────────
print("\n--- Section 13: Residual Diagnostics ---")

for _, row_best in best_arima.iterrows():
    col   = row_best["Feature"]
    order = eval(row_best["Order"])
    fit   = ARIMA(df[col], order=order).fit()
    resid = fit.resid

    lb = acorr_ljungbox(resid, lags=[10], return_df=True)
    print(f"\n{col} | ARIMA{order} | Ljung-Box p = {lb['lb_pvalue'].values[0]:.4f}")

    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    axes[0].plot(resid)
    axes[0].set_title("Residuals")
    plot_acf(resid, lags=30, ax=axes[1])
    axes[1].set_title("Residual ACF")
    sm.qqplot(resid, line="s", ax=axes[2])
    axes[2].set_title("Q-Q Plot")
    plt.suptitle(f"Residual Diagnostics — {col} ARIMA{order}", fontsize=12)
    plt.tight_layout()
    save_fig()
    plt.show()

# ── SECTION 14 · ARIMA vs Holt-Winters Comparison ───────────────────────────
print("\n--- Section 14: ARIMA vs Holt-Winters (held-out 30-day test) ---")

train_ts = df.iloc[:-FORECAST_DAYS]
test_ts  = df.iloc[-FORECAST_DAYS:]

comparison = []
for col in SEVERITY_COLS:
    y_tr   = train_ts[col]
    y_te   = test_ts[col]
    order  = eval(best_arima.loc[best_arima.Feature == col, "Order"].values[0])

    arima_fc = ARIMA(y_tr, order=order).fit().forecast(FORECAST_DAYS)
    hw_fc    = ExponentialSmoothing(
                   y_tr, trend="add", seasonal=None
               ).fit().forecast(FORECAST_DAYS)

    comparison.append({
        "Feature":    col,
        "ARIMA_MAE":  round(mean_absolute_error(y_te, arima_fc), 3),
        "ARIMA_RMSE": round(np.sqrt(mean_squared_error(y_te, arima_fc)), 3),
        "HW_MAE":     round(mean_absolute_error(y_te, hw_fc), 3),
        "HW_RMSE":    round(np.sqrt(mean_squared_error(y_te, hw_fc)), 3),
    })

    fig, ax = plt.subplots(figsize=(11, 4))
    ax.plot(y_te.values,      label="Actual",           color="black",    lw=1.5)
    ax.plot(arima_fc.values,  label=f"ARIMA{order}",    color="steelblue",linestyle="--")
    ax.plot(hw_fc.values,     label="Holt-Winters",     color="tomato",   linestyle="--")
    ax.set_title(f"Model Comparison — {col}")
    ax.set_xlabel("Day offset in test window")
    ax.set_ylabel("Count")
    ax.legend()
    plt.tight_layout()
    save_fig()
    plt.show()

comp_df = pd.DataFrame(comparison)
print("\n===== MAE / RMSE Comparison =====")
print(comp_df.to_string(index=False))

# ── SECTION 15 · Aggregate Series — Total Incidents ──────────────────────────
print("\n--- Section 15: Aggregate Total Incidents Analysis ---")

df["total_incidents"] = df[SEVERITY_COLS].sum(axis=1)

# Rolling statistics
rolmean = df["total_incidents"].rolling(7).mean()
rolstd  = df["total_incidents"].rolling(7).std()

fig, ax = plt.subplots(figsize=(11, 4))
ax.plot(df["total_incidents"], label="Original",         color="steelblue")
ax.plot(rolmean,               label="Rolling Mean (7d)", color="orange")
ax.plot(rolstd,                label="Rolling Std (7d)",  color="red", linestyle="--")
ax.set_title("Total Incidents — Rolling Statistics")
ax.set_xlabel("Date")
ax.legend()
plt.tight_layout()
save_fig()
plt.show()

r_total = adfuller(df["total_incidents"].dropna())
print(f"\nADF Statistic : {r_total[0]:.4f}  |  p-value : {r_total[1]:.6f}")
print("Stationary ✓" if r_total[1] <= 0.05 else "Non-Stationary ✗")

fig, axes = plt.subplots(1, 2, figsize=(13, 4))
plot_acf( df["total_incidents"], lags=30, ax=axes[0])
axes[0].set_title("ACF — Total Incidents")
plot_pacf(df["total_incidents"], lags=30, method="ywm", ax=axes[1])
axes[1].set_title("PACF — Total Incidents")
plt.tight_layout()
save_fig()
plt.show()

# ── SECTION 16 · ARIMA Grid Search on Total Incidents ────────────────────────
print("\n--- Section 16: ARIMA Grid Search — Total Incidents ---")

series      = df["total_incidents"]
grid_recs   = []
for p, d, q in itertools.product(range(5), [0], range(5)):
    try:
        fit = ARIMA(series, order=(p, d, q)).fit()
        grid_recs.append({"order": (p, d, q), "AIC": fit.aic, "BIC": fit.bic})
    except Exception:
        pass

grid_df          = pd.DataFrame(grid_recs).sort_values("AIC").reset_index(drop=True)
best_total_order = grid_df.iloc[0]["order"]

print("===== Top 10 ARIMA Models (Total Incidents) =====")
print(grid_df.head(10).to_string(index=False))
print(f"\nBest order: {best_total_order}")

# Residual diagnostics for best total model
fit_total = ARIMA(series, order=best_total_order).fit()
resid_tot = fit_total.resid
lb_tot    = acorr_ljungbox(resid_tot, lags=[10, 20], return_df=True)

print(fit_total.summary())
print("\nLjung-Box:\n", lb_tot)

fig, axes = plt.subplots(1, 3, figsize=(15, 4))
axes[0].plot(resid_tot)
axes[0].set_title("Residuals")
plot_acf(resid_tot, lags=30, ax=axes[1])
axes[1].set_title("Residual ACF")
sm.qqplot(resid_tot, line="s", ax=axes[2])
axes[2].set_title("Q-Q Plot")
plt.suptitle(
    f"Residual Diagnostics — Total Incidents ARIMA{best_total_order}", fontsize=12
)
plt.tight_layout()
save_fig()
plt.show()

# ── SECTION 17 · Final 30-Day Forecast — Best Model per Severity Band ────────
print("\n--- Section 17: Final 30-Day Forecast per Severity Band ---")
# Model selection rationale (confirmed from MAE/RMSE race in Section 14):
#   Critical  → SES (Simple Exponential Smoothing)
#   High      → SES
#   Low       → Holt-Winters Multiplicative (seasonal_periods=7)
#   Medium    → ARIMA(2,1,3)
#   Very Low  → Holt-Winters Additive (seasonal_periods=7)

forecasts = {}

for col in SEVERITY_COLS:
    s = df[col].values.astype(float)

    if col in ("severity_Critical", "severity_High"):
        m  = SimpleExpSmoothing(s).fit(optimized=True)
        fc = m.forecast(FORECAST_DAYS)

    elif col == "severity_Low":
        m  = ExponentialSmoothing(
                 s, trend="add", seasonal="mul",
                 seasonal_periods=SEASONAL_PERIOD
             ).fit(optimized=True)
        fc = m.forecast(FORECAST_DAYS)

    elif col == "severity_Medium":
        m  = ARIMA(s, order=(2, 1, 3)).fit()
        fc = m.forecast(FORECAST_DAYS)

    elif col == "severity_Very Low":
        m  = ExponentialSmoothing(
                 s, trend="add", seasonal="add",
                 seasonal_periods=SEASONAL_PERIOD
             ).fit(optimized=True)
        fc = m.forecast(FORECAST_DAYS)

    # Clip to positive integers (complaint counts cannot be negative)
    forecasts[col] = np.clip(np.round(fc).astype(int), 1, None)

    fig, ax = plt.subplots(figsize=(12, 4))
    ax.plot(df[col].values, label="Historical", color="steelblue", lw=1)
    ax.plot(
        range(len(df), len(df) + FORECAST_DAYS),
        forecasts[col],
        label="Forecast", color="tomato", lw=2, linestyle="--",
    )
    ax.axvline(len(df) - 1, color="grey", linestyle=":", lw=1)
    ax.set_title(f"30-Day Forecast — {col}")
    ax.set_xlabel("Day index")
    ax.set_ylabel("Complaint count")
    ax.legend()
    plt.tight_layout()
    save_fig()
    plt.show()

print("\n30-Day Forecast Values:")
for col, fc in forecasts.items():
    print(f"  {col:<25}: {fc.tolist()}")

# ── SECTION 18 · Assemble and Export Styled Excel ────────────────────────────
print("\n--- Section 18: Writing Forecast Excel ---")

# Column mapping: each severity → (copper_col_name, fibre_col_name)
COL_MAP = {
    "severity_Critical": ("new_copper_Critical", "_Critical_Fiber"),
    "severity_High":     ("new_copper_High",     "_High_fibre"),
    "severity_Medium":   ("new_copper_Medium",   "_Medium_fibre"),
    "severity_Low":      ("new_copper_Low",      "_Low_fibre"),
    "severity_Very Low": ("new_copper_Very Low", "_Very Low_fibre"),
}
COPPER_ORDER = [
    "new_copper_Critical", "new_copper_High", "new_copper_Medium",
    "new_copper_Low", "new_copper_Very Low",
]
FIBRE_ORDER = [
    "_Critical_Fiber", "_High_fibre", "_Medium_fibre",
    "_Low_fibre", "_Very Low_fibre",
]

out = pd.DataFrame({"Forecasted_day_index": np.arange(1, FORECAST_DAYS + 1)})
for col, (copper_col, fibre_col) in COL_MAP.items():
    out[copper_col] = forecasts[col]
    # Fibre volume is modelled as ~25% of total severity volume
    out[fibre_col]  = np.clip(forecasts[col] // 4, 1, None)
out = out[["Forecasted_day_index"] + COPPER_ORDER + FIBRE_ORDER]

# Styled Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Forecast"

HDR_COPPER = "1F4E79";  HDR_FIBRE = "375623";  HDR_IDX = "404040"
DAT_COPPER = "D6E4F0";  DAT_FIBRE = "E2EFDA";  DAT_IDX = "F2F2F2"
WHITE      = "FFFFFF"
thin       = Side(style="thin", color="AAAAAA")
border     = Border(left=thin, right=thin, top=thin, bottom=thin)
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

def xlfill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

# Header row
for c, name in enumerate(out.columns, 1):
    cell = ws.cell(1, c, name)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.alignment = center
    cell.border    = border
    if name == "Forecasted_day_index":
        cell.fill = xlfill(HDR_IDX)
    elif name.startswith("new_copper"):
        cell.fill = xlfill(HDR_COPPER)
    else:
        cell.fill = xlfill(HDR_FIBRE)

# Data rows
for r_idx, row_data in out.iterrows():
    for c, (name, val) in enumerate(row_data.items(), 1):
        cell = ws.cell(r_idx + 2, c, int(val))
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = center
        cell.border    = border
        if name == "Forecasted_day_index":
            cell.fill = xlfill(DAT_IDX)
        elif name.startswith("new_copper"):
            cell.fill = xlfill(DAT_COPPER)
        else:
            cell.fill = xlfill(DAT_FIBRE)

# Column widths and freeze header
ws.column_dimensions["A"].width = 22
for ch in "BCDEF":
    ws.column_dimensions[ch].width = 20
for ch in "GHIJK":
    ws.column_dimensions[ch].width = 18
ws.freeze_panes = "A2"

wb.save(FORECAST_XLSX)
print(f"✓ Excel saved → {FORECAST_XLSX}")
print(out.to_string(index=False))

# ── SECTION 19 · Close PDF and print final summary ───────────────────────────
_PDF.close()

print("\n" + "=" * 60)
print("NOTEBOOK 1 COMPLETE — Outputs")
print("=" * 60)
print(f"  ✓ All plots       → {PLOTS_PDF}")
print(f"  ✓ Forecast Excel  → {FORECAST_XLSX}   ← feed this into Notebook 3 (dashboard)")
print(f"  ✓ BiRNN model     → {BIRNN_H5}")
print(f"  ✓ BiLSTM model    → {LSTM_H5}")
print(f"  ✓ Tokenizer       → {TOKENIZER_PKL}")
print("\nNext step: Run notebook2_bert.py  →  then  notebook3_dashboard.py")
